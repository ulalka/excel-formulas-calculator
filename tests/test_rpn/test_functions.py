# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, print_function, unicode_literals

from itertools import chain, groupby, product
from os.path import dirname, join

import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from efc.interfaces.iopenpyxl import OpenpyxlInterface
from efc.rpn_builder.parser.operands import (
    BadReference,
    NotFoundErrorOperand, NumErrorOperand,
    ValueErrorOperand,
    ValueNotAvailable,
    ZeroDivisionErrorOperand,
)
from .mock import ExcelMock, get_calculator


@pytest.fixture(scope='session')
def calc():
    source = ExcelMock()
    calculator = get_calculator()
    return lambda line, ws_name: calculator(line, ws_name, source)


@pytest.fixture(scope='session')
def workbook():
    path = join(dirname(__file__), 'fixtures', 'functions.xlsx')
    return load_workbook(path)


@pytest.fixture(scope='session')
def workbook_data_only():
    path = join(dirname(__file__), 'fixtures', 'functions.xlsx')
    return load_workbook(path, data_only=True)


@pytest.fixture(scope='session')
def interface(workbook):
    return OpenpyxlInterface(workbook, use_cache=True)


def test_SUM(calc):
    assert calc('SUM(Sheet4!A1:B3)', 'Yet another sheet').value == 64
    assert calc('SUM([0]Sheet4!A1:B3)', 'Yet another sheet').value == 64
    assert calc('SUM(Sheet4!A1:B3) + 1', 'Yet another sheet').value == 65
    assert calc('SUM(Sheet4!A1:B3,A2:B3)', 'Sheet4').value == 99
    assert calc('SUM(Sheet4!A1:B3,SUM(A3:B3))', 'Sheet4').value == 70


def test_SUMIFS(calc):
    assert calc('SUMIFS(Sheet4!A1:B3,Sheet4!A1:B3,">4")', 'Yet another sheet').value == 58
    assert calc('SUMIFS(Sheet4!A1:B3,Sheet4!A1:B3,"13")', 'Yet another sheet').value == 26
    assert calc('SUMIFS(Sheet4!A1:B3,Sheet4!A1:C5,"<=0")', 'Yet another sheet').value == 0
    assert calc(
        'SUMIFS(Sheet5!A1:C5,Sheet5!A1:C5,"<>0",Sheet5!B2:D4,">0")', 'Yet another sheet').value == 29


def test_SUMIF(calc):
    assert calc('SUMIF(Sheet4!A1:B3,">4")', 'Yet another sheet').value == 58
    assert calc('SUMIF(Sheet4!A1:B3,">4",Sheet4!A1:B3)', 'Yet another sheet').value == 58
    assert calc('SUMIF(Sheet4!A1:B3,"13")', 'Yet another sheet').value == 26
    assert calc('SUMIF(Sheet4!A1:B3,"13",Sheet4!B2:C4)', 'Yet another sheet').value == 18
    assert calc('SUMIF(Sheet4!A1:B5,"=0")', 'Yet another sheet').value == 0
    assert calc('SUMIF(Sheet5!A1:C5,">=0")', 'Yet another sheet').value == 49


def test_MOD(calc):
    assert calc('MOD(\'Sheet 1\'!B3,4)', 'Yet another sheet').value == 2
    assert calc('MOD(\'Sheet 1\'!A3,\'Sheet 1\'!C3)', 'Yet another sheet').value == 4
    assert calc('MOD(\'Sheet 1\'!A3,\'Sheet 1\'!B3 * 2)', 'Yet another sheet').value == 0


def test_IF(calc):
    assert calc('IF(2>1,1,2)', 'Yet another sheet').value == 1
    assert calc('IF(TRUE,1,2)', 'Yet another sheet').value == 1
    assert calc('IF(FALSE,1,2)', 'Yet another sheet').value == 2
    assert calc('IF(TRUE,1,2 ** 5)', 'Yet another sheet').value == 1
    assert calc('IF(\'Sheet 1\'!A3 = 4,\'Sheet 1\'!C3, 0)', 'Yet another sheet').value == 8

    assert calc('IF(TRUE, 1,)', 'Yet another sheet').value == 1
    assert calc('IF(FALSE, 1,)', 'Yet another sheet').value == 0
    assert calc('CONCATENATE(IF(FALSE, 1,))', 'Yet another sheet').value == ''


@pytest.mark.parametrize('prefix', ('_xlfn.', '_xludf.', ''))
def test_IFS(calc, prefix):
    assert calc(prefix + 'IFS(FALSE,1,FALSE,2,TRUE,3)', 'Yet another sheet').value == 3
    assert calc(prefix + 'IFS(TRUE,1,FALSE,2,TRUE,3)', 'Yet another sheet').value == 1
    assert calc(prefix + 'IFS(FALSE,1,TRUE,2,TRUE,3)', 'Yet another sheet').value == 2

    with pytest.raises(ValueNotAvailable):
        assert calc(prefix + 'IFS(FALSE,1,FALSE,2,FALSE,3)', 'Yet another sheet').value == 1


def test_IFERROR(calc):
    assert calc('IFERROR(5/0,1)', 'Yet another sheet').value == 1
    assert calc('IFERROR(5+6, 0)', 'Yet another sheet').value == 11


def test_MAX(calc):
    assert calc('MAX(Sheet4!A1:B3)', 'Yet another sheet').value == 16
    assert calc('MAX(Sheet4!A1:B3,100)', 'Yet another sheet').value == 100


def test_MIN(calc):
    assert calc('MIN(Sheet4!A1:B3)', 'Yet another sheet').value == 2
    assert calc('MIN(Sheet4!A1:B3,1)', 'Yet another sheet').value == 1


def test_LEFT(calc):
    assert calc('LEFT("test", 2)', 'Yet another sheet').value == 'te'


def test_RIGHT(calc):
    assert calc('RIGHT("test", 2)', 'Yet another sheet').value == 'st'


def test_MID(calc):
    assert calc('MID("hello",2,2)', 'Sheet 1').value == 'el'


@pytest.mark.parametrize(('f', 'result'),
                         (('ISBLANK("test")', False),
                          ('ISBLANK("")', False),
                          ('ISBLANK(Sheet4!AA1)', True),
                          ))
def test_ISBLANK(calc, f, result):
    assert calc(f, 'Yet another sheet').value is result


def test_OR(calc):
    assert calc('OR(0,0,0,TRUE)', 'Yet another sheet').value is True
    assert calc('OR(FALSE, 0)', 'Yet another sheet').value is False
    assert calc('OR(FALSE, 0 + 2)', 'Yet another sheet').value is True


def test_AND(calc):
    assert calc('AND(1,1,1,TRUE)', 'Yet another sheet').value is True
    assert calc('AND(FALSE, 0)', 'Yet another sheet').value is False
    assert calc('AND(TRUE, 0 + 2)', 'Yet another sheet').value is True


def test_not(calc):
    assert calc('NOT(1)', 'Yet another sheet').value is False
    assert calc('NOT(0)', 'Yet another sheet').value is True
    assert calc('NOT(123)', 'Yet another sheet').value is False
    assert calc('NOT(TRUE)', 'Yet another sheet').value is False
    assert calc('NOT(FALSE)', 'Yet another sheet').value is True
    assert calc('NOT(A1)', 'Yet another sheet').value is True
    assert calc('NOT(A1)', 'Sheet4').value is False
    assert calc('NOT("")', 'Yet another sheet').value is True
    assert calc('NOT(NOT(""))', 'Yet another sheet').value is False


def test_ROUND(calc):
    assert calc('ROUND(2.3456, 1)', 'Yet another sheet').value == 2.3
    assert calc('ROUND(2, 2)', 'Yet another sheet').value == 2.0
    assert calc('ROUND("2.34567", 2)', 'Yet another sheet').value == 2.35


def test_ROUNDDOWN(calc):
    assert calc('ROUNDDOWN(1.345,0)', 'Sheet 1').value == 1.0
    assert calc('ROUNDDOWN(1.345,1)', 'Sheet 1').value == 1.3
    assert calc('ROUNDDOWN(1.345,2)', 'Sheet 1').value == 1.34


def test_FLOOR(calc):
    assert calc('FLOOR(10,3)', 'Sheet 1').value == 9
    assert calc('FLOOR(16,7)', 'Sheet 1').value == 14
    assert calc('FLOOR(26,13)', 'Sheet 1').value == 26


def test_COUNT(calc):
    assert calc('COUNT(1.3456, 1, "test")', 'Yet another sheet').value == 2
    assert calc('COUNT(A1:C4)', 'Sheet 1').value == 6


@pytest.mark.parametrize(('f', 'sheet', 'result'),
                         (
                                 ('COUNTIF(A1:C4, ">4")', 'Sheet 1', 4),
                                 ('COUNTIF(A1:C4, "13")', 'Sheet4', 2),
                                 ('COUNTIF(A1:C5, ">=0")', 'Sheet4', 9),
                                 ('COUNTIF(A1:C5, "<=0")', 'Sheet5', 0),
                                 ('COUNTIF(A1:B5, "")', 'Sheet6', 2),
                                 ('COUNTIF(A1:B5, B4)', 'Sheet6', 3),
                                 ('COUNTIF(A1:B5, 0)', 'Sheet6', 3),
                                 ('COUNTIF(A1:B5, A1)', 'Sheet6', 3),
                                 ('COUNTIF(A1:B5, C1)', 'Sheet6', 2),
                         ))
def test_simple_COUNTIF(calc, f, sheet, result):
    assert calc(f, sheet).value == result


def test_COUNTBLANK(calc):
    assert calc('COUNTBLANK(A1:C4)', 'Sheet 1').value == 6
    assert calc('COUNTBLANK(A1:B4)', 'Sheet4').value == 2


def test_ABS(calc):
    assert calc('ABS(1.32)', 'Sheet 1').value == 1.32
    assert calc('ABS(-42)', 'Sheet4').value == 42


def test_OFFSET(calc):
    assert calc('OFFSET(A1,2,1)', 'Sheet 1').value == 2
    assert calc('OFFSET(A1,2,1,1)', 'Sheet 1').value == 2
    assert calc('OFFSET(A1,B3,1)', 'Sheet 1').value == 2
    assert calc('SUM(OFFSET(A1,2,1,1,2))', 'Sheet 1').value == 10


def test_simple_MATCH(calc):
    assert calc('MATCH(13,Sheet4!A1:A3)', 'Yet another sheet').value == 2


def test_AVERAGE(calc):
    assert calc('AVERAGE(Sheet4!A1:B3)', 'Yet another sheet').value == 64 / 6
    assert calc('AVERAGEIFS(Sheet4!A1:B3,Sheet4!A1:B3,"13")', 'Yet another sheet').value == 13


def test_AVERAGEIFS(calc):
    assert calc('AVERAGEIFS(Sheet4!A1:B3,Sheet4!A1:B3,"13")', 'Yet another sheet').value == 13
    assert calc('AVERAGEIFS(Sheet4!A1:C5,Sheet4!A1:C5,"<>0")', 'Yet another sheet').value == 12
    assert calc('AVERAGEIFS(Sheet5!A1:C5,Sheet5!A2:C6,"<>0",Sheet5!B2:D6,"<18")', 'Yet another sheet').value == 13


@pytest.mark.parametrize(
    ('formula', 'result'),
    (('VLOOKUP(13,TestVLookup!A1:B3,2)', 17),
     ('VLOOKUP(13,TestVLookup!A1:B3,2,1)', 17),
     ('VLOOKUP(13,TestVLookup!A1:B3,2,0)', 16),
     ('VLOOKUP(4,\'Yet another sheet\'!A100:ZZ110,27)', 45),
     pytest.param('VLOOKUP(4,TestVLookup!A1:B3,2)', 12435,
                  marks=pytest.mark.xfail(raises=NotFoundErrorOperand, strict=True)),
     pytest.param('VLOOKUP(24,TestVLookup!A1:B3,2)', 12435,
                  marks=pytest.mark.xfail(raises=NotFoundErrorOperand, strict=True))),
)
def test_VLOOKUP(calc, formula, result):
    assert calc(formula, 'Sheet4').value == result


@pytest.mark.parametrize(
    ('formula', 'result'),
    (('HLOOKUP(13,TestHLookup!A1:B3,2)', 17),
     ('HLOOKUP(13,TestHLookup!A1:B3,2,1)', 17),
     ('HLOOKUP(13,TestHLookup!A1:B3,2,0)', 16),
     ('HLOOKUP(45,\'Yet another sheet\'!A104:ZZ04,1)', 45),
     pytest.param('HLOOKUP(4,TestHLookup!A1:B3,2)', 12435,
                  marks=pytest.mark.xfail(raises=NotFoundErrorOperand, strict=True)),
     pytest.param('HLOOKUP(24,TestHLookup!A1:B3,2)', 12435,
                  marks=pytest.mark.xfail(raises=NotFoundErrorOperand, strict=True))),
)
def test_HLOOKUP(calc, formula, result):
    assert calc(formula, 'Sheet4').value == result


def test_SEARCH(calc):
    assert calc('SEARCH("abc", "abc")', 'Yet another sheet').value == 1
    assert calc('SEARCH("abc", "abc", 1)', 'Yet another sheet').value == 1
    assert calc('SEARCH("abc", "aabc", 1)', 'Yet another sheet').value == 2

    with pytest.raises(ValueErrorOperand):
        assert calc('SEARCH("abcd", "abc")', 'Yet another sheet').value

    with pytest.raises(ValueErrorOperand):
        assert calc('SEARCH("abcd", "abc", 1)', 'Yet another sheet').value


def test_SMALL(calc):
    assert calc('SMALL(Sheet4!A1:B3,1)', 'Yet another sheet').value == 2
    assert calc('SMALL(Sheet4!A1:B3,2)', 'Yet another sheet').value == 4
    assert calc('SMALL(Sheet4!A1:B3,4)', 'Yet another sheet').value == 13


def test_LARGE(calc):
    assert calc('LARGE(Sheet4!A1:B3,1)', 'Yet another sheet').value == 16
    assert calc('LARGE(Sheet4!A1:B3,2)', 'Yet another sheet').value == 16
    assert calc('LARGE(Sheet4!A1:B3,4)', 'Yet another sheet').value == 13


def test_COUNTIFS(calc):
    assert calc('COUNTIFS(Sheet4!A1:B3,">4")', 'Yet another sheet').value == 4
    assert calc('COUNTIFS(Sheet4!A1:B3,"13")', 'Yet another sheet').value == 2
    assert calc('COUNTIFS(Sheet4!A1:B4,"<>0")', 'Yet another sheet').value == 8
    assert calc('COUNTIFS(Sheet5!A1:C5,"=0")', 'Yet another sheet').value == 0


def test_COUNTA(calc):
    assert calc('COUNTA(Sheet4!A1:B4)', 'Yet another sheet').value == 6
    assert calc('COUNTA(Sheet5!A1:B4)', 'Yet another sheet').value == 5


def test_CONCATENATE(calc):
    assert calc('CONCATENATE(Sheet4!A1,Sheet4!B3,"13")', 'Yet another sheet').value == '13213'
    assert calc('CONCATENATE("",Sheet4!B3,TRUE)', 'Yet another sheet').value == '2TRUE'


def test_INDEX(calc):
    assert calc('INDEX(Sheet4!A1:A3,1)', 'Yet another sheet').value == 13  # A1
    assert calc('INDEX(Sheet4!A1:B3,1,2)', 'Yet another sheet').value == 16  # B1
    assert calc('INDEX(Sheet4!A1:A3,3)', 'Yet another sheet').value == 4  # A3
    assert calc('INDEX(Sheet4!A2:A3,2)', 'Yet another sheet').value == 4  # A3

    assert calc('INDEX(Sheet4!A:A,3)', 'Yet another sheet').value == 4  # A3
    assert calc('INDEX(Sheet4!A:A,3,1)', 'Yet another sheet').value == 4  # A3
    assert calc('INDEX(Sheet4!3:3,1,1)', 'Yet another sheet').value == 4  # A3
    assert calc('INDEX(Sheet4!3:3,1,2)', 'Yet another sheet').value == 2  # B3

    with pytest.raises(BadReference):
        assert calc('INDEX(Sheet4!A1:A3,100,1)', 'Yet another sheet').value

    with pytest.raises(BadReference):
        assert calc('INDEX(Sheet4!A1:A3,1,100)', 'Yet another sheet').value

    assert calc('INDEX(Sheet4!A1:C3,1,1)', 'Yet another sheet').value == 13
    assert calc('INDEX(Sheet4!A1:C3,1,3)', 'Yet another sheet').value == 18

    with pytest.raises(BadReference):
        assert calc('INDEX(Sheet4!A1:C3,1)', 'Yet another sheet').value

    with pytest.raises(BadReference):
        assert calc('INDEX(Sheet4!A1:C3,100,1)', 'Yet another sheet').value

    with pytest.raises(BadReference):
        assert calc('INDEX(Sheet4!A1:C3,1,100)', 'Yet another sheet').value

    with pytest.raises(BadReference):
        assert calc('INDEX(Sheet4!A1:C3,0,100)', 'Yet another sheet').value

    with pytest.raises(BadReference):
        assert calc('INDEX(Sheet4!A1:C3,1,0)', 'Yet another sheet').value

    with pytest.raises(BadReference):
        assert calc('INDEX(Sheet4!A:A,3,2)', 'Yet another sheet').value

    with pytest.raises(BadReference):
        assert calc('INDEX(Sheet4!1:1,2)', 'Yet another sheet').value


def test_SUBSTITUTE(calc):
    assert calc('SUBSTITUTE("123123123","1","22")', 'Yet another sheet').value == '222322232223'
    assert calc('SUBSTITUTE("123123123","1","22", 2)', 'Yet another sheet').value == '22232223123'
    assert calc('SUBSTITUTE("123123123","1","22", -1)', 'Yet another sheet').value == '222322232223'


def test_TRIM(calc):
    assert calc('TRIM(1)', 'Yet another sheet').value == '1'
    assert calc('TRIM(0)', 'Yet another sheet').value == '0'
    assert calc('TRIM("1")', 'Yet another sheet').value == '1'
    assert calc('TRIM(" 1 ")', 'Yet another sheet').value == '1'
    assert calc('TRIM(" 1 1 ")', 'Yet another sheet').value == '1 1'
    assert calc('TRIM(" 1      1 ")', 'Yet another sheet').value == '1 1'


def test_LEN(calc):
    assert calc('LEN(1)', 'Yet another sheet').value == 1
    assert calc('LEN(0)', 'Yet another sheet').value == 1
    assert calc('LEN("1")', 'Yet another sheet').value == 1
    assert calc('LEN(" 1 ")', 'Yet another sheet').value == 3


def test_YEARFRAC(calc):
    with pytest.raises(NumErrorOperand):
        assert calc('YEARFRAC(1, 2, 5)', 'Yet another sheet').value

    # 30U/360
    assert calc('YEARFRAC(43159, 43160)', 'Yet another sheet').value == 1 / 360
    assert calc('YEARFRAC(43405, 43465, 0)', 'Yet another sheet').value == 60 / 360
    assert calc('YEARFRAC(43889, 43890, 0)', 'Yet another sheet').value == 1 / 360
    assert calc('YEARFRAC(43889, 43891, 0)', 'Yet another sheet').value == 3 / 360

    # Actual/Actual
    assert calc('YEARFRAC(43889, 43890, 1)', 'Yet another sheet').value == 1 / 366
    assert calc('YEARFRAC(43889, 43891, 1)', 'Yet another sheet').value == 2 / 366
    assert calc('YEARFRAC(43523, 43524, 1)', 'Yet another sheet').value == 1 / 365
    assert calc('YEARFRAC(43523, 43525, 1)', 'Yet another sheet').value == 2 / 365

    # Actual/360
    assert calc('YEARFRAC(43889, 43890, 2)', 'Yet another sheet').value == 1 / 360
    assert calc('YEARFRAC(43889, 43891, 2)', 'Yet another sheet').value == 2 / 360
    assert calc('YEARFRAC(43523, 43524, 2)', 'Yet another sheet').value == 1 / 360
    assert calc('YEARFRAC(43523, 43525, 2)', 'Yet another sheet').value == 2 / 360

    # Actual/365
    assert calc('YEARFRAC(43889, 43890, 3)', 'Yet another sheet').value == 1 / 365
    assert calc('YEARFRAC(43889, 43891, 3)', 'Yet another sheet').value == 2 / 365
    assert calc('YEARFRAC(43523, 43524, 3)', 'Yet another sheet').value == 1 / 365
    assert calc('YEARFRAC(43523, 43525, 3)', 'Yet another sheet').value == 2 / 365

    # 30E/360
    assert calc('YEARFRAC(43159, 43160, 4)', 'Yet another sheet').value == 3 / 360
    assert calc('YEARFRAC(43405, 43465, 4)', 'Yet another sheet').value == 59 / 360
    assert calc('YEARFRAC(43889, 43890, 4)', 'Yet another sheet').value == 1 / 360
    assert calc('YEARFRAC(43889, 43891, 4)', 'Yet another sheet').value == 3 / 360


def test_lower(calc):
    with pytest.raises(ZeroDivisionErrorOperand):
        _ = calc('LOWER(1/0)', 'Yet another sheet').value

    assert calc('LOWER(1)', 'Yet another sheet').value == '1'
    assert calc('LOWER("1")', 'Yet another sheet').value == '1'
    assert calc('LOWER("TEST")', 'Yet another sheet').value == 'test'
    assert calc('LOWER("test")', 'Yet another sheet').value == 'test'
    assert calc('LOWER("TeSt %1234")', 'Yet another sheet').value == 'test %1234'


def test_upper(calc):
    with pytest.raises(ZeroDivisionErrorOperand):
        _ = calc('UPPER(1/0)', 'Yet another sheet').value

    assert calc('UPPER(1)', 'Yet another sheet').value == '1'
    assert calc('UPPER("1")', 'Yet another sheet').value == '1'
    assert calc('UPPER("TEST")', 'Yet another sheet').value == 'TEST'
    assert calc('UPPER("test")', 'Yet another sheet').value == 'TEST'
    assert calc('UPPER("TeSt %1234")', 'Yet another sheet').value == 'TEST %1234'


@pytest.mark.parametrize(
    ('rng', 'by_col', 'exactly_once', 'result'),
    (
            ('A1:D1', None, None, (('A', 'B', 'B', 'B'),)),
            ('A1:D1', True, None, (('A', 'B'),)),
            ('A1:D1', False, None, (('A', 'B', 'B', 'B'),)),
            ('A1:D1', True, None, (('A', 'B'),)),
            ('A1:D1', True, True, (('A',),)),
            ('A1:D1', False, False, (('A', 'B', 'B', 'B'),)),
            ('A1:D1', False, True, (('A', 'B', 'B', 'B'),)),
            ('A1:A10', None, None, (('A',), (1,), (2,), (3,), (None,), (5,), (7,), ('7',), (0,),)),
            ('A1:A10', True, None, (('A',), (1,), (2,), (3,), (None,), (5,), (7,), (7,), ('7',), (0,),)),
            ('A1:A10', False, None, (('A',), (1,), (2,), (3,), (None,), (5,), (7,), ('7',), (0,),)),
            ('A1:A10', True, None, (('A',), (1,), (2,), (3,), (None,), (5,), (7,), (7,), ('7',), (0,),)),
            ('A1:A10', True, True, (('A',), (1,), (2,), (3,), (None,), (5,), (7,), (7,), ('7',), (0,),)),
            ('A1:A10', False, False, (('A',), (1,), (2,), (3,), (None,), (5,), (7,), ('7',), (0,),)),
            ('A1:A10', False, True, (('A',), (1,), (2,), (3,), (None,), (5,), ('7',), (0,),)),
            ('A1:D10', None, None, (('A', 'B', 'B', 'B'),
                                    (1, 1, 6, 6),
                                    (2, 2, 7, 7),
                                    (3, 3, 8, 8),
                                    (None, None, 9, 9),
                                    (5, 5, 10, 10),
                                    (7, 7, 6, 6),
                                    ('7', 7, 6, 6),
                                    (0, 0, 9, 9),
                                    )),
            ('A1:D10', True, None, (('A', 'B', 'B'),
                                    (1, 1, 6),
                                    (2, 2, 7),
                                    (3, 3, 8),
                                    (None, None, 9),
                                    (5, 5, 10,),
                                    (7, 7, 6),
                                    (7, 7, 6),
                                    ('7', 7, 6),
                                    (0, 0, 9),
                                    )),
            ('A1:D10', False, None, (('A', 'B', 'B', 'B'),
                                     (1, 1, 6, 6),
                                     (2, 2, 7, 7),
                                     (3, 3, 8, 8),
                                     (None, None, 9, 9),
                                     (5, 5, 10, 10),
                                     (7, 7, 6, 6),
                                     ('7', 7, 6, 6),
                                     (0, 0, 9, 9),
                                     )),
            ('A1:D10', True, None, (('A', 'B', 'B'),
                                    (1, 1, 6),
                                    (2, 2, 7),
                                    (3, 3, 8),
                                    (None, None, 9),
                                    (5, 5, 10,),
                                    (7, 7, 6),
                                    (7, 7, 6),
                                    ('7', 7, 6),
                                    (0, 0, 9),
                                    )),
            ('A1:D10', True, True, (('A', 'B'),
                                    (1, 1),
                                    (2, 2),
                                    (3, 3),
                                    (None, None),
                                    (5, 5,),
                                    (7, 7),
                                    (7, 7),
                                    ('7', 7),
                                    (0, 0),
                                    )),
            ('A1:D10', False, False, (('A', 'B', 'B', 'B'),
                                      (1, 1, 6, 6),
                                      (2, 2, 7, 7),
                                      (3, 3, 8, 8),
                                      (None, None, 9, 9),
                                      (5, 5, 10, 10),
                                      (7, 7, 6, 6),
                                      ('7', 7, 6, 6),
                                      (0, 0, 9, 9),
                                      )),
            ('A1:D10', False, True, (('A', 'B', 'B', 'B'),
                                     (1, 1, 6, 6),
                                     (2, 2, 7, 7),
                                     (3, 3, 8, 8),
                                     (None, None, 9, 9),
                                     (5, 5, 10, 10),
                                     ('7', 7, 6, 6),
                                     (0, 0, 9, 9),
                                     )),
    )
)
def test_unique(calc, rng, by_col, exactly_once, result):
    suffix = ''
    if by_col is not None:
        suffix += (',' + str(by_col).upper())
    if exactly_once:
        suffix += (',' + str(exactly_once).upper())
    op = calc('UNIQUE(TestUnique!%s%s)' % (rng, suffix), 'TestUnique')
    op_result = []
    for _, cells in groupby(op.get_iter_rows(), lambda x: x[0]):
        op_result.append(tuple(c.value for _, c in cells))

    assert tuple(op_result) == result


@pytest.mark.parametrize(
    ('formula', 'ws', 'result'),
    (('ROW(A1)', 'Yet another sheet', 1),
     ('ROW(C1)', 'Yet another sheet', 1),
     ('ROW(C3)', 'Yet another sheet', 3),
     ('ROW(GH12435)', 'Yet another sheet', 12435),
     pytest.param('ROW(1234)', 'Yet another sheet', 12435,
                  marks=pytest.mark.xfail(raises=ValueErrorOperand, strict=True)),
     pytest.param('ROW("123")', 'Yet another sheet', 12435,
                  marks=pytest.mark.xfail(raises=ValueErrorOperand, strict=True))),
)
def test_row(calc, formula, ws, result):
    assert calc(formula, ws).value == result


@pytest.mark.parametrize(
    ('formula', 'ws', 'result'),
    (('COLUMN(A1)', 'Yet another sheet', 1),
     ('COLUMN(C1)', 'Yet another sheet', 3),
     ('COLUMN(GH12435)', 'Yet another sheet', 190),
     pytest.param('COLUMN(1234)', 'Yet another sheet', 12435,
                  marks=pytest.mark.xfail(raises=ValueErrorOperand, strict=True)),
     pytest.param('COLUMN("123")', 'Yet another sheet', 12435,
                  marks=pytest.mark.xfail(raises=ValueErrorOperand, strict=True))),
)
def test_column(calc, formula, ws, result):
    assert calc(formula, ws).value == result


COUNTIF_COMPARE_VALUES = (
    '-100',
    '-1',
    '0',
    'zero_link',
    'empty_string_link',
    'empty_cell_link',
    'empty_string_formula',
    '1',
    '100',
    'int_string_formula',
    'int_string_link',
    'str_string_formula',
    'str_string_link',
    'string',
    'str_link',
)

COUNTIF_OPERANDS = ('<>', '>=', '<=', '>', '<', '=', '= (wo sign)')
COUNTIF_FIRST_DATA_ROW = 2
COUNTIF_FIRST_OPERAND_COLUMN = 6
COUNTIF_LIST = 'countif'


@pytest.mark.parametrize('compare_value', enumerate(COUNTIF_COMPARE_VALUES))
@pytest.mark.parametrize('operand', enumerate(COUNTIF_OPERANDS))
def test_countif(workbook_data_only, workbook, interface, compare_value, operand):
    compare_value_idx, compare_value = compare_value
    operand_idx, operand_value = operand
    row = COUNTIF_FIRST_DATA_ROW + compare_value_idx
    column = COUNTIF_FIRST_OPERAND_COLUMN + operand_idx
    index = get_column_letter(column) + str(row)

    # TODO get value and formula together from single fixture
    v = workbook_data_only[COUNTIF_LIST][index].value
    expected = int(v) if v != 'skip' else v
    assert interface.calc_cell(index, COUNTIF_LIST) == expected, index


MATCH_VALUES = COUNTIF_COMPARE_VALUES
MATCH_FIRST_OPERAND_COLUMN = 6
MATCH_LIST = 'match'

MATCH_INT_FIRST_DATA_ROW = 3
MATCH_INT_VALUES = (
    '-100',
    '-1',
    '0',
    'zero_link',
    'empty_cell_link',
    '1',
    '100',
)

MATCH_STRING_FIRST_DATA_ROW = 24
MATCH_STRING_VALUES = (
    'empty_string_link',
    'empty_cell_link',
    'empty_string_formula',
    'int_string_formula',
    'int_string_link',
    'str_string_formula',
    'str_string_link',
    'string',
    'str_link',
)


@pytest.mark.parametrize('match_type', enumerate((1, 0)))
@pytest.mark.parametrize(('first_row', 'data_type', 'operand'),
                         chain(
                             product((MATCH_INT_FIRST_DATA_ROW,), ('int',), list(enumerate(MATCH_INT_VALUES))),
                             product((MATCH_STRING_FIRST_DATA_ROW,), ('string',), list(enumerate(MATCH_STRING_VALUES))),
                         ))
def test_match(workbook_data_only, workbook, interface, first_row, data_type, operand, match_type):
    operand_idx, operand_value = operand
    row = first_row + operand_idx
    column = MATCH_FIRST_OPERAND_COLUMN + match_type[0]
    index = get_column_letter(column) + str(row)

    # TODO get value and formula together from single fixture
    expected = workbook_data_only[MATCH_LIST][index].value
    assert interface.calc_cell(index, MATCH_LIST) == expected, index
