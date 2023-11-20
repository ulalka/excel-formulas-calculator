# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, print_function, unicode_literals

from os.path import dirname, join

import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from efc.interfaces.iopenpyxl import OpenpyxlInterface
from tests.test_rpn.mock import get_calculator


@pytest.fixture(scope='module')
def calculate():
    calculator = get_calculator()
    return lambda line: calculator(line, None, None)


@pytest.fixture(scope='session')
def workbook():
    path = join(dirname(__file__), 'fixtures', 'arith.xlsx')
    return load_workbook(path)


@pytest.fixture(scope='session')
def workbook_data_only():
    path = join(dirname(__file__), 'fixtures', 'arith.xlsx')
    return load_workbook(path, data_only=True)


@pytest.fixture(scope='session')
def interface(workbook):
    return OpenpyxlInterface(workbook, use_cache=True)


@pytest.mark.parametrize(
    ['line', 'result'],
    (
            ('4', 4),
            ('-4', -4),
            ('4 + 4', 8),
            ('-4 + 4', 0),
            ('4 - 2', 2),
            ('4 * 4', 16),
            ('9 ^ 2', 81),
            ('9 / 3', 3),
            ('1 + 2 * 3', 7),
            ('2 * 3 + 1', 7),
            ('2 - (2 - 3)', 3),
            ('2 - (-2 - 3)', 7),
            ('2 - 2 * 8 - 3', -17),
            ('2 - 2 - 3 - 6', -9),
            ('2 - 2 - 3', -3),
            ('2 - 2 + 3', 3),
            ('2 - (2 + 3)', -3),
    )
)
def test_arithmetic(calculate, line, result):
    assert calculate(line).value == result


@pytest.mark.parametrize(
    ['line', 'result'],
    (
            ('4 = 3', False),
            ('4 <> 3', True),
            ('4 > 3', True),
            ('4 >= 4', True),
            ('4 < 3', False),
            ('3 <= 3', True),
            ('4 + 1 > 4', True),
            ('4 > 4 - 3', True),
            ('4 * 2 + 2 <> 4 / 3 - 1', True),

            ('1 > TRUE', False),
            ('1 > FALSE', False),
            ('1 > "1"', False),
            ('"1" > 1', True),

            ('1 = TRUE', False),
            ('1 = FALSE', False),
            ('1 = "1"', False),
            ('"1" = 1', False),

            ('1 < TRUE', True),
            ('1 < FALSE', True),
            ('1 < "1"', True),
            ('"1" < 1', False),
    )
)
def test_simple_compare(calculate, line, result):
    assert calculate(line).value == result


@pytest.mark.parametrize(
    ['line', 'result'],
    (
            ('"test" & "2"', 'test2'),
            ('2 & 3', '23'),
    )
)
def test_concat(calculate, line, result):
    assert calculate(line).value == result


COMPARE_VALUES = (
    'zero',
    'zero_link',
    'empty_string_formula',
    'empty_string_link',
    'empty_cell_link',
    'zero_link_link',
    'empty_string_link_link',
    'empty_cell_link_link',
)

COMPARE_OPERANDS = ('<>', '>=', '<=', '>', '<', '=')
FIRST_DATA_ROW = 2
FIRST_OPERAND_COLUMN = 5
COMPARE_LIST = 'compare'


@pytest.mark.parametrize('left', enumerate(COMPARE_VALUES))
@pytest.mark.parametrize('right', enumerate(COMPARE_VALUES))
@pytest.mark.parametrize('operand', enumerate(COMPARE_OPERANDS))
def test_compare(workbook_data_only, workbook, interface, left, right, operand):
    left_idx, left_value = left
    right_idx, right_value = right
    operand_idx, operand_value = operand
    row = FIRST_DATA_ROW + left_idx + len(COMPARE_VALUES) * right_idx
    column = FIRST_OPERAND_COLUMN + operand_idx
    index = get_column_letter(column) + str(row)

    # TODO get value and formula together from single fixture
    expected = workbook_data_only[COMPARE_LIST][index].value
    assert interface.calc_cell(index, COMPARE_LIST) == expected, \
        '%s %s %s != %s' % (left_value, operand_value, right_value, expected)
